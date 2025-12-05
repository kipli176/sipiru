import io
import csv
from flask import session
from openpyxl import Workbook, load_workbook 
from db import query_all, query_one, execute

# =====================================================
# ===============  UTILITIES / GENERAL  ===============
# ===================================================== 

def q(sql, params=None, one=False, commit=False, many=False):
    if many:
        execute(sql, params, many=True)
        return None
    if one:
        return query_one(sql, params)
    if commit:
        execute(sql, params)
        return None
    return query_all(sql, params)

def to_str(v):
    return None if v is None else str(v)

def get_total(table, where_sql, params):
    sql = f"SELECT COUNT(*) AS total FROM {table} WHERE {where_sql}"
    return q(sql, params, one=True)["total"]

def get_all_timeslots():
    return q("SELECT * FROM time_slot ORDER BY id")

def parse_slot_labels(slot_labels):
    if not slot_labels:
        return "", ""
    first_label = slot_labels[0]
    last_label = slot_labels[-1]
    jam_mulai = first_label.split("-")[0].strip() if "-" in first_label else ""
    jam_selesai = last_label.split("-")[1].strip() if "-" in last_label else ""
    return jam_mulai, jam_selesai

def parse_import_file(file, min_cols=1):
    filename = file.filename.lower()
    rows = []

    if filename.endswith(".csv"):
        content = file.stream.read().decode("utf-8")
        reader = csv.reader(content.splitlines())
        for r in reader:
            clean = [str(c).strip() for c in r]
            if len(clean) >= min_cols:
                rows.append(clean)

    elif filename.endswith(".xlsx"):
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            clean = [(str(v).strip() if v else "") for v in row]
            if len(clean) >= min_cols:
                rows.append(clean)

    else:
        return None, "Format file harus CSV atau XLSX"

    return rows, None

def generate_excel(sheet_name, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)

    for r in rows:
        ws.append(r)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =====================================================
# ==========  AUTH / ADMIN & FAKULTAS HELPERS =========
# =====================================================

def get_fakultas_aktif_nama():
    return q("SELECT id, nama FROM fakultas WHERE aktif=TRUE ORDER BY nama")

def authenticate_admin(fakultas_id, password):
    return q("""
        SELECT *
        FROM admin_user
        WHERE (fakultas_id=%s OR role='superadmin')
          AND pin_hash=%s
          AND aktif=TRUE
        LIMIT 1
    """, [fakultas_id, password], one=True)

def set_admin_session(admin):
    session["admin_id"] = admin["id"]
    session["admin_nama"] = admin["nama"]
    session["role"] = admin["role"]
    session["fakultas_id"] = admin["fakultas_id"]

    fak = q("SELECT nama FROM fakultas WHERE id=%s", [admin["fakultas_id"]], one=True)
    session["fakultas_nama"] = fak["nama"] if fak else None


# =====================================================
# ================= DASHBOARD HELPERS =================
# =====================================================

def get_dashboard_counts():
    total_booking = q("SELECT COUNT(*) AS total FROM v_booking_group", one=True)["total"]
    total_ruangan = q("""
        SELECT COUNT(*) AS total
        FROM v_ruangan_fakultas
        WHERE aktif=TRUE
    """, one=True)["total"]
    total_jadwal = q("SELECT COUNT(*) AS total FROM jadwal_kuliah", one=True)["total"]

    return total_booking, total_ruangan, total_jadwal

def get_dashboard_recent_activity():
    rows = q("""
        SELECT *
        FROM v_booking_group
        ORDER BY created_at DESC
        LIMIT 5
    """)

    formatted = []
    for r in rows:
        jam_mulai, jam_selesai = parse_slot_labels(r["slot_labels"])
        formatted.append({
            **r,
            "jam_mulai": jam_mulai,
            "jam_selesai": jam_selesai
        })
    return formatted

def get_top_ruangan():
    return q("""
        SELECT nama_ruang, kode_ruang, COUNT(*) AS jumlah
        FROM v_booking_group
        GROUP BY nama_ruang, kode_ruang
        ORDER BY jumlah DESC
        LIMIT 3
    """)

def get_jadwal_hari_ini(today):
    return q("""
        SELECT *
        FROM v_jadwal_ruang_harian
        WHERE tanggal=%s AND source='kuliah'
        ORDER BY time_slot_id
    """, [today])


# =====================================================
# =================== RUANGAN HELPERS =================
# =====================================================

def get_ruangan_page(fak, per_page, offset):
    return q("""
        SELECT *
        FROM ruangan
        WHERE fakultas_id=%s
        ORDER BY kode_ruang
        LIMIT %s OFFSET %s
    """, [fak, per_page, offset])

def get_ruangan_by_id(id, fak):
    return q("SELECT * FROM ruangan WHERE id=%s AND fakultas_id=%s",
             [id, fak], one=True)

def get_ruangan_list(fak):
    return q("""
        SELECT id, kode_ruang, nama_ruang
        FROM ruangan
        WHERE fakultas_id=%s
        ORDER BY kode_ruang
    """, [fak])

def ruang_kode_exists(fak, kode, exclude_id=None):
    if exclude_id:
        return q("""
            SELECT 1 FROM ruangan
            WHERE fakultas_id=%s AND kode_ruang=%s AND id!=%s
        """, [fak, kode, exclude_id], one=True)
    return q("""
        SELECT 1 FROM ruangan
        WHERE fakultas_id=%s AND kode_ruang=%s
    """, [fak, kode], one=True)

def insert_ruangan(fak, kode, nama, gedung, lantai, kapasitas, fasilitas, catatan):
    q("""
        INSERT INTO ruangan (
            fakultas_id, kode_ruang, nama_ruang, gedung,
            lantai, kapasitas, fasilitas, aktif, catatan
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
    """, [fak, kode, nama, gedung, lantai, kapasitas, fasilitas, catatan], commit=True)

def update_ruangan(id, fak, kode, nama, gedung, lantai, kapasitas, fasilitas, aktif, catatan):
    q("""
        UPDATE ruangan
        SET kode_ruang=%s, nama_ruang=%s, gedung=%s,
            lantai=%s, kapasitas=%s, fasilitas=%s,
            aktif=%s, catatan=%s
        WHERE id=%s AND fakultas_id=%s
    """, [kode, nama, gedung, lantai, kapasitas, fasilitas, aktif, catatan, id, fak], commit=True)

def toggle_ruangan(id, fak):
    q("UPDATE ruangan SET aktif = NOT aktif WHERE id=%s AND fakultas_id=%s",
      [id, fak], commit=True)

def delete_ruangan(id, fak):
    q("DELETE FROM ruangan WHERE id=%s AND fakultas_id=%s",
      [id, fak], commit=True)

def get_ruangan_export_rows(fak):
    return q("""
        SELECT kode_ruang, nama_ruang, gedung, lantai,
               kapasitas, fasilitas, catatan
        FROM ruangan
        WHERE fakultas_id = %s
        ORDER BY kode_ruang
    """, [fak])

def generate_ruangan_export_excel(rows):
    header = ["kode_ruang", "nama_ruang", "gedung", "lantai",
              "kapasitas", "fasilitas", "catatan"]
    formatted = [[
        r["kode_ruang"], r["nama_ruang"], r["gedung"], r["lantai"],
        r["kapasitas"], r["fasilitas"], r["catatan"]
    ] for r in rows]
    return generate_excel("Ruangan", header, formatted)


# =====================================================
# ==================== JADWAL HELPERS =================
# =====================================================

def get_jadwal_page(fak, per_page, offset):
    return q("""
        SELECT jk.*, r.kode_ruang
        FROM jadwal_kuliah jk
        LEFT JOIN ruangan r ON r.id = jk.ruangan_id
        WHERE jk.fakultas_id=%s
        ORDER BY jk.hari_dow ASC, jk.jam_mulai ASC, jk.id DESC
        LIMIT %s OFFSET %s
    """, [fak, per_page, offset])

def get_jadwal_detail(id, fak):
    return q("""
        SELECT jk.*, r.kode_ruang, r.nama_ruang
        FROM jadwal_kuliah jk
        LEFT JOIN ruangan r ON r.id = jk.ruangan_id
        WHERE jk.id=%s AND jk.fakultas_id=%s
    """, [id, fak], one=True)

def bentrok_jadwal(ruangan_id, hari_dow, jam_mulai, jam_selesai):
    return q("""
        SELECT 1
        FROM jadwal_kuliah
        WHERE ruangan_id=%s
          AND hari_dow=%s
          AND NOT (jam_selesai <= %s OR jam_mulai >= %s)
    """, [ruangan_id, hari_dow, jam_mulai, jam_selesai], one=True)

def insert_jadwal(data):
    q("""
        INSERT INTO jadwal_kuliah (
            semester_kode, tanggal_mulai, tanggal_selesai,
            hari_dow, hari_nama, jam_mulai, jam_selesai,
            ruangan_id, mata_kuliah, kode_mk, kelas, dosen,
            fakultas_id
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, data, commit=True)

def update_jadwal(id, fak, form):
    q("""
        UPDATE jadwal_kuliah
        SET semester_kode=%s,
            tanggal_mulai=%s, tanggal_selesai=%s,
            hari_dow=%s, hari_nama=%s,
            jam_mulai=%s, jam_selesai=%s,
            ruangan_id=%s, mata_kuliah=%s,
            kode_mk=%s, kelas=%s, dosen=%s
        WHERE id=%s AND fakultas_id=%s
    """, [
        form["semester_kode"],
        form.get("tanggal_mulai"),
        form.get("tanggal_selesai"),
        form["hari_dow"], form["hari"],
        form["jam_mulai"], form["jam_selesai"],
        form["ruangan_id"], form["mata_kuliah"],
        form.get("kode_mk"), form.get("kelas"), form.get("dosen"),
        id, fak
    ], commit=True)

def delete_jadwal(id, fak):
    q("DELETE FROM jadwal_kuliah WHERE id=%s AND fakultas_id=%s",
      [id, fak], commit=True)

def batch_insert_jadwal(batch_values):
    sql = """
        INSERT INTO jadwal_kuliah (
            hari_nama, jam_mulai, jam_selesai, ruangan_id,
            mata_kuliah, kelas, dosen,
            semester_kode, kode_mk,
            tanggal_mulai, tanggal_selesai,
            hari_dow, fakultas_id
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """
    q(sql, batch_values, many=True, commit=True)

def get_jadwal_export_rows(fak):
    return q("""
        SELECT jk.*, r.kode_ruang
        FROM jadwal_kuliah jk
        LEFT JOIN ruangan r ON r.id = jk.ruangan_id
        WHERE jk.fakultas_id=%s
        ORDER BY jk.hari_dow ASC, jk.jam_mulai ASC
    """, [fak])

def generate_jadwal_export_excel(rows):
    header = [
        "hari","jam_mulai","jam_selesai","kode_ruang",
        "mata_kuliah","kelas","dosen","semester_kode",
        "kode_mk","tanggal_mulai","tanggal_selesai"
    ]
    formatted = [[
        r["hari_nama"], r["jam_mulai"], r["jam_selesai"],
        r["kode_ruang"], r["mata_kuliah"], r["kelas"],
        r["dosen"], r["semester_kode"], r["kode_mk"],
        r["tanggal_mulai"], r["tanggal_selesai"]
    ] for r in rows]
    return generate_excel("Jadwal", header, formatted)


# =====================================================
# ==================== BOOKING HELPERS ================
# =====================================================

def get_booking_page(fak, per_page, offset):
    return q("""
        SELECT *
        FROM v_booking_group
        WHERE fakultas_ruangan_id=%s
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """, [fak, per_page, offset])

def serialize_booking_rows(rows):
    data = []
    for r in rows:
        jam_mulai, jam_selesai = parse_slot_labels(r["slot_labels"])
        data.append({
            "id": r["any_booking_id"],
            "group_id": r["group_id"],
            "tanggal": str(r["tanggal"]),
            "kode_ruang": r["kode_ruang"],
            "jam_mulai": jam_mulai,
            "jam_selesai": jam_selesai,
            "nama_peminjam": r["nama_peminjam"],
            "kegiatan": r["kegiatan"],
            "status": r["status"],
            "created_at": str(r["created_at"]),
        })
    return data
